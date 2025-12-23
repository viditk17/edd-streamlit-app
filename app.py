import io
import os
import re
import shutil
import tempfile
from datetime import time
from typing import Optional, Tuple, List

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# =============================================================================
# ✅ OPENAI SDK (ROBUST IMPORT: works with new + legacy, avoids Streamlit crash)
# =============================================================================
# - New SDK (recommended): openai>=1.x => from openai import OpenAI
# - Legacy SDK: openai==0.28.1 => import openai; openai.ChatCompletion.create
try:
    from openai import OpenAI  # type: ignore
except Exception:
    OpenAI = None  # type: ignore

try:
    import openai as openai_legacy  # type: ignore
except Exception:
    openai_legacy = None  # type: ignore


# =============================================================================
# ✅ PROCESSOR (YOUR FULL BACKEND CODE) — Summary + NDR + Grouping + Formatting
# =============================================================================

def process_edd_report(
    input_path: str,
    output_path: Optional[str] = None,
    source_sheet: str = "Query result",
    summary_sheet: str = "summary",
    apply_formatting: bool = True,
) -> str:
    """
    Process EDD Excel and return the processed file path.

    - If output_path is None, input file is overwritten.
    - Otherwise input file is copied to output_path first, and output is written there.
    """
    if output_path is None:
        file_path = input_path
    else:
        shutil.copyfile(input_path, output_path)
        file_path = output_path

    # ===================== CONFIG =====================
    FILE_PATH = file_path
    SOURCE_SHEET = source_sheet
    SUMMARY_SHEET = summary_sheet

    # ===================== LOAD DATA =====================
    df = pd.read_excel(FILE_PATH, sheet_name=SOURCE_SHEET)
    df.columns = df.columns.str.strip()

    date_cols = ["EDD_Date", "PICKUP_CHLN_DATE", "Reached At Destination", "DLY_Date"]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    df = df.dropna(subset=["EDD_Date"])

    # ===================== NEW EDD =====================
    df["NEW_EDD_DATE"] = df["PICKUP_CHLN_DATE"] + pd.to_timedelta(
        df["TAT_DAYS"] - 1, unit="D"
    )

    # ===================== ON TIME ARRIVAL =====================
    df["ON_TIME_ARRIVAL"] = "No"
    valid_arrival = (
        df["Reached At Destination"].notna()
        & df["EDD_Date"].notna()
        & df["NEW_EDD_DATE"].notna()
    )

    cond1 = df["Reached At Destination"] <= df["NEW_EDD_DATE"]
    cond2 = (
        (df["Reached At Destination"].dt.date == df["EDD_Date"].dt.date)
        & (df["Reached At Destination"].dt.time < time(12, 0))
    )

    df.loc[valid_arrival & (cond1 | cond2), "ON_TIME_ARRIVAL"] = "Yes"

    # ===================== ON TIME DELIVERY =====================
    df["ON_TIME_DELIVERY"] = "No"
    df.loc[
        df["DLY_Date"].notna()
        & df["EDD_Date"].notna()
        & (df["DLY_Date"] <= df["EDD_Date"]),
        "ON_TIME_DELIVERY",
    ] = "Yes"

    # ===================== WEEK LABEL =====================
    df["Week_Label"] = "W-" + df["EDD_Date"].dt.isocalendar().week.astype(str)

    # ===================== TOTALS =====================
    weekly_total = df.groupby("Week_Label").size().reset_index(name="Picked Volume")

    zone_week = (
        df.groupby(["Week_Label", "BKG_Zone"]).size().reset_index(name="Zone_CN")
    )
    zone_week = zone_week.merge(weekly_total, on="Week_Label", how="left")
    zone_week["Zone_Percent"] = (
        zone_week["Zone_CN"] / zone_week["Picked Volume"] * 100
    ).round(2)

    mode_counts = (
        df.groupby(["Week_Label", "BKG_Zone", "TPTR_Mode"])
        .size()
        .reset_index(name="Mode_CN")
    )

    # ===================== BUILD SUMMARY =====================
    weeks = weekly_total["Week_Label"].tolist()
    zones = sorted(df["BKG_Zone"].dropna().unique())
    all_modes = sorted(df["TPTR_Mode"].dropna().unique())
    all_statuses = sorted(df["CN_Current_Status"].dropna().unique())
    all_business_types = sorted(df["BUSINESS_TYPE"].dropna().unique())

    summary = pd.DataFrame(columns=weeks)
    summary.loc["Picked Volume"] = weekly_total.set_index("Week_Label")["Picked Volume"]

    # ===================== ZONE BLOCK =====================
    for zone in zones:
        zdf = zone_week[zone_week["BKG_Zone"] == zone]

        summary.loc[f"Picked Vol. Zone {zone} %"] = {
            wk: (
                f"{r['Zone_Percent'].values[0]}% ({int(r['Zone_CN'].values[0])})"
                if not (r := zdf[zdf["Week_Label"] == wk]).empty
                else "0% (0)"
            )
            for wk in weeks
        }

        # ===================== BUSINESS TYPE BREAKDOWN =====================
        summary.loc[f"BUSINESS TYPE BREAKDOWN__{zone}"] = {wk: "" for wk in weeks}

        for bt in all_business_types:
            summary.loc[f"   {bt}__{zone}"] = {
                wk: (
                    f"{round((cnt / tc.values[0]) * 100, 2)}% ({cnt})"
                    if not tc.empty and tc.values[0] != 0
                    else "0% (0)"
                )
                for wk, cnt, tc in [
                    (
                        wk,
                        df[
                            (df["BKG_Zone"] == zone)
                            & (df["Week_Label"] == wk)
                            & (df["BUSINESS_TYPE"] == bt)
                        ].shape[0],
                        zdf[zdf["Week_Label"] == wk]["Zone_CN"],
                    )
                    for wk in weeks
                ]
            }

        # ===================== TPTR MODE =====================
        for mode in all_modes:
            mdf = mode_counts[
                (mode_counts["BKG_Zone"] == zone) & (mode_counts["TPTR_Mode"] == mode)
            ]

            mode_total = {
                wk: int(mdf[mdf["Week_Label"] == wk]["Mode_CN"].values[0])
                if not mdf[mdf["Week_Label"] == wk].empty
                else 0
                for wk in weeks
            }

            summary.loc[f"TPTR Mode {mode}__{zone}"] = {
                wk: (
                    f"{round((mode_total[wk] / tc.values[0]) * 100, 2)}% ({mode_total[wk]})"
                    if mode_total[wk] > 0 and not tc.empty
                    else "0% (0)"
                )
                for wk, tc in {wk: zdf[zdf["Week_Label"] == wk]["Zone_CN"] for wk in weeks}.items()
            }

            summary.loc[f"{mode} On Time Arrival__{zone}"] = {
                wk: (
                    f"{round((ota / mode_total[wk]) * 100, 2)}% ({ota})"
                    if mode_total[wk] > 0
                    else "0% (0)"
                )
                for wk, ota in {
                    wk: df[
                        (df["BKG_Zone"] == zone)
                        & (df["Week_Label"] == wk)
                        & (df["TPTR_Mode"] == mode)
                        & (df["ON_TIME_ARRIVAL"] == "Yes")
                    ].shape[0]
                    for wk in weeks
                }.items()
            }

            summary.loc[f"{mode} On Time Delivery__{zone}"] = {
                wk: (
                    f"{round((otd / mode_total[wk]) * 100, 2)}% ({otd})"
                    if mode_total[wk] > 0
                    else "0% (0)"
                )
                for wk, otd in {
                    wk: df[
                        (df["BKG_Zone"] == zone)
                        & (df["Week_Label"] == wk)
                        & (df["TPTR_Mode"] == mode)
                        & (df["ON_TIME_DELIVERY"] == "Yes")
                    ].shape[0]
                    for wk in weeks
                }.items()
            }

        # ===================== NDR NOT AVAILABLE (YOUR RULE) =====================
        # CN_Current_Status == "Ware house Destination" AND NDR_Remark is blank
        summary.loc[f"NDR not available__{zone}"] = {
            wk: (
                f"{round((cnt / tc.values[0]) * 100, 2)}% ({cnt})"
                if not tc.empty and tc.values[0] != 0
                else "0% (0)"
            )
            for wk, cnt, tc in [
                (
                    wk,
                    df[
                        (df["BKG_Zone"] == zone)
                        & (df["Week_Label"] == wk)
                        & (df["CN_Current_Status"] == "Ware house Destination")
                        & (
                            df["NDR_Remark"].isna()
                            | (df["NDR_Remark"].astype(str).str.strip() == "")
                        )
                    ].shape[0],
                    zdf[zdf["Week_Label"] == wk]["Zone_CN"],
                )
                for wk in weeks
            ]
        }

        # ===================== CN STATUS BREAKDOWN =====================
        summary.loc[f"CN Status Breakdown__{zone}"] = {wk: "" for wk in weeks}

        for status in all_statuses:
            summary.loc[f"   {status}__{zone}"] = {
                wk: (
                    f"{round((cnt / tc.values[0]) * 100, 2)}% ({cnt})"
                    if not tc.empty and tc.values[0] != 0
                    else "0% (0)"
                )
                for wk, cnt, tc in [
                    (
                        wk,
                        df[
                            (df["BKG_Zone"] == zone)
                            & (df["Week_Label"] == wk)
                            & (df["CN_Current_Status"] == status)
                        ].shape[0],
                        zdf[zdf["Week_Label"] == wk]["Zone_CN"],
                    )
                    for wk in weeks
                ]
            }

    # ===================== CLEAN LABELS =====================
    summary.index = summary.index.str.replace(r"__(.*)$", "", regex=True)

    # ===================== WRITE SUMMARY =====================
    with pd.ExcelWriter(
        FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        summary.to_excel(writer, sheet_name=SUMMARY_SHEET)

    # ===================== APPLY ROW GROUPING (FULL HIERARCHY) =====================
    wb = load_workbook(FILE_PATH)
    ws = wb[SUMMARY_SHEET]

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    max_row = ws.max_row
    row = 2

    while row <= max_row:
        val = ws[f"A{row}"].value

        # ===================== ZONE LEVEL =====================
        if isinstance(val, str) and val.startswith("Picked Vol. Zone"):
            ws.row_dimensions[row].outline_level = 1
            ws.row_dimensions[row].collapsed = True

            zone_start = row + 1
            r = zone_start
            while r <= max_row and not (
                isinstance(ws[f"A{r}"].value, str)
                and ws[f"A{r}"].value.startswith("Picked Vol. Zone")
            ):
                r += 1
            zone_end = r - 1

            # Hide everything under zone
            if zone_end >= zone_start:
                ws.row_dimensions.group(zone_start, zone_end, hidden=True, outline_level=2)

            sub = zone_start
            while sub <= zone_end:
                txt = ws[f"A{sub}"].value

                # ===================== BUSINESS TYPE =====================
                if txt == "BUSINESS TYPE BREAKDOWN":
                    ws.row_dimensions[sub].outline_level = 2
                    ws.row_dimensions[sub].collapsed = True

                    s = sub + 1
                    e = s
                    while (
                        e <= zone_end
                        and isinstance(ws[f"A{e}"].value, str)
                        and ws[f"A{e}"].value.startswith("   ")
                    ):
                        e += 1

                    if e - 1 >= s:
                        ws.row_dimensions.group(s, e - 1, hidden=True, outline_level=3)

                    sub = e
                    continue

                # ===================== TPTR MODE =====================
                if isinstance(txt, str) and txt.startswith("TPTR Mode"):
                    ws.row_dimensions[sub].outline_level = 2
                    ws.row_dimensions[sub].collapsed = True

                    s = sub + 1
                    e = s
                    while e <= zone_end and not (
                        isinstance(ws[f"A{e}"].value, str)
                        and (
                            ws[f"A{e}"].value.startswith("TPTR Mode")
                            or ws[f"A{e}"].value == "CN Status Breakdown"
                            or ws[f"A{e}"].value == "BUSINESS TYPE BREAKDOWN"
                            or ws[f"A{e}"].value.startswith("NDR")
                        )
                    ):
                        e += 1

                    if e - 1 >= s:
                        ws.row_dimensions.group(s, e - 1, hidden=True, outline_level=3)

                    sub = e
                    continue

                # ===================== CN STATUS =====================
                if txt == "CN Status Breakdown":
                    ws.row_dimensions[sub].outline_level = 2
                    ws.row_dimensions[sub].collapsed = True

                    s = sub + 1
                    e = s
                    while (
                        e <= zone_end
                        and isinstance(ws[f"A{e}"].value, str)
                        and ws[f"A{e}"].value.startswith("   ")
                    ):
                        e += 1

                    if e - 1 >= s:
                        ws.row_dimensions.group(s, e - 1, hidden=True, outline_level=3)

                    sub = e
                    continue

                sub += 1

            row = r
        else:
            row += 1

    wb.save(FILE_PATH)

    if apply_formatting:
        _format_summary_sheet(FILE_PATH, summary_sheet=SUMMARY_SHEET)

    return FILE_PATH


def _format_summary_sheet(file_path: str, summary_sheet: str = "summary") -> None:
    FILE_PATH = file_path
    SHEET_NAME = summary_sheet

    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_NAME]

    # --- Styles ---
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="2F2F2F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Freeze panes ---
    ws.freeze_panes = "B2"

    # --- Auto width (simple) ---
    ws.column_dimensions["A"].width = 42
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

    # --- Header row style ---
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # --- First column bold ---
    for r in range(2, ws.max_row + 1):
        c = ws[f"A{r}"]
        c.font = bold_font
        c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Borders for all cells ---
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center

    # --- Conditional Formatting ---
    start_row = 2
    start_col = 2
    end_row = ws.max_row
    end_col = ws.max_column
    rng = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(end_row, end_col).coordinate}"

    rule = ColorScaleRule(
        start_type="min",
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="63BE7B",
    )
    ws.conditional_formatting.add(rng, rule)

    wb.save(FILE_PATH)


# =============================================================================
# ✅ OPENAI + STREAMLIT (CHAT + PROCESS) — FIXED (NO CRASH ON CLOUD)
# =============================================================================

def get_openai_key() -> Optional[str]:
    # Priority: secrets.toml -> env var -> sidebar input
    key = None
    try:
        key = st.secrets.get("OPENAI_API_KEY", None)
    except Exception:
        key = None
    if not key:
        key = os.getenv("OPENAI_API_KEY") or os.getenv("OPENAI_SECRET_KEY")
    if not key:
        key = st.session_state.get("_openai_key")
    return key


def make_client() -> Optional[object]:
    """
    Returns:
      - OpenAI(api_key=...) client if new SDK available
      - openai_legacy module if legacy SDK available
      - None if no key or no SDK installed
    """
    key = get_openai_key()
    if not key:
        return None

    if OpenAI is not None:
        return OpenAI(api_key=key)

    if openai_legacy is not None and hasattr(openai_legacy, "ChatCompletion"):
        openai_legacy.api_key = key
        return openai_legacy

    return None


def llm_answer(client: object, model: str, system: str, user: str) -> str:
    """
    - New SDK: prefers Responses API w/ instructions=system, input=user
    - Fallback: Chat Completions
    - Legacy SDK fallback: openai.ChatCompletion.create
    Never raises (won't crash Streamlit); returns a readable error string on failure.
    """
    last_err: Optional[Exception] = None

    # ---- New SDK: Responses API ----
    if hasattr(client, "responses"):
        try:
            resp = client.responses.create(
                model=model,
                instructions=system,
                input=user,
            )
            out = getattr(resp, "output_text", None)
            if out:
                return out.strip()
            return str(resp)
        except Exception as e:
            last_err = e

    # ---- New SDK: Chat Completions ----
    if hasattr(client, "chat") and hasattr(getattr(client, "chat"), "completions"):
        try:
            comp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                temperature=0.2,
            )
            return (comp.choices[0].message.content or "").strip()
        except Exception as e:
            last_err = e

    # ---- Legacy SDK: ChatCompletion ----
    if hasattr(client, "ChatCompletion"):
        try:
            comp = client.ChatCompletion.create(
                model=model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                temperature=0.2,
            )
            # legacy returns dict-like
            return (comp["choices"][0]["message"]["content"] or "").strip()
        except Exception as e:
            last_err = e

    return f"❌ OpenAI call failed: {last_err}" if last_err else "❌ OpenAI client not available."


@st.cache_data(show_spinner=False)
def load_summary_df(xlsx_bytes: bytes, sheet: str = "summary") -> pd.DataFrame:
    bio = io.BytesIO(xlsx_bytes)
    df = pd.read_excel(bio, sheet_name=sheet, engine="openpyxl")
    if df.columns.size > 0 and str(df.columns[0]).lower().startswith("unnamed"):
        df = df.rename(columns={df.columns[0]: "Metric"})
    return df


def _normalize_week_label(week_num: int) -> str:
    return f"W-{int(week_num)}"


def extract_week(question: str) -> Optional[str]:
    """
    Accepts:
      - week 48, week-48, week48
      - w 48, w-48, w48
      - W-48 (already)
    Returns: 'W-48' or None
    """
    q = (question or "").strip()

    m = re.search(r"\bW\s*[-_ ]\s*(\d{1,2})\b", q, flags=re.IGNORECASE)
    if m:
        return _normalize_week_label(int(m.group(1)))

    m = re.search(r"\bweek\s*[-_ ]*\s*(\d{1,2})\b", q, flags=re.IGNORECASE)
    if m:
        return _normalize_week_label(int(m.group(1)))

    m = re.search(r"\bweek(\d{1,2})\b", q, flags=re.IGNORECASE)
    if m:
        return _normalize_week_label(int(m.group(1)))

    return None


def zones_from_summary(summary_df: pd.DataFrame) -> List[str]:
    if "Metric" not in summary_df.columns:
        return []
    zones = []
    for v in summary_df["Metric"].astype(str).tolist():
        m = re.match(r"^Picked Vol\. Zone (.*) %$", v.strip())
        if m:
            zones.append(m.group(1).strip())
    return sorted(set(zones))


def extract_zone(question: str, zones: List[str]) -> Optional[str]:
    q = (question or "").upper()
    zones_sorted = sorted(zones, key=lambda z: len(z), reverse=True)
    for z in zones_sorted:
        if z.upper() in q:
            return z
    return None


def select_relevant_rows(question: str, summary_df: pd.DataFrame) -> List[str]:
    q = (question or "").lower()
    metric_col = "Metric" if "Metric" in summary_df.columns else summary_df.columns[0]
    metrics = summary_df[metric_col].astype(str).tolist()

    hits = []

    # basic intents
    if "picked" in q and "volume" in q:
        if "Picked Volume" in metrics:
            hits.append("Picked Volume")

    if "cn" in q and ("status" in q or "current" in q):
        if "CN Status Breakdown" in metrics:
            hits.append("CN Status Breakdown")

    if "ndr" in q:
        if "NDR not available" in metrics:
            hits.append("NDR not available")

    if "business" in q or "retail" in q or "scm" in q:
        if "BUSINESS TYPE BREAKDOWN" in metrics:
            hits.append("BUSINESS TYPE BREAKDOWN")

    if "tptr" in q or "mode" in q:
        for m in metrics:
            if str(m).startswith("TPTR Mode"):
                hits.append(m)
            if ("arrival" in q or "ota" in q) and "On Time Arrival" in str(m):
                hits.append(m)
            if ("delivery" in q or "otd" in q) and "On Time Delivery" in str(m):
                hits.append(m)

    # fallback keyword matching
    if not hits:
        tokens = [t for t in re.findall(r"[a-zA-Z0-9]+", q) if len(t) > 2]
        stop = {
            "the", "and", "with", "from", "this", "that",
            "mein", "me", "ka", "ki", "ke", "for",
            "show", "dikhao", "bata", "batao", "please"
        }
        tokens = [t for t in tokens if t not in stop]
        for m in metrics:
            mm = m.lower()
            if any(t in mm for t in tokens):
                hits.append(m)

    seen = set()
    out = []
    for h in hits:
        if h not in seen and h in metrics:
            seen.add(h)
            out.append(h)
    return out


def answer_from_summary(
    summary_df: pd.DataFrame, question: str
) -> Tuple[Optional[str], Optional[str], List[Tuple[str, str]]]:
    if summary_df.empty:
        return None, None, []

    metric_col = "Metric" if "Metric" in summary_df.columns else summary_df.columns[0]

    zones = zones_from_summary(summary_df)
    week = extract_week(question)

    # If week missing, use latest available week col
    week_cols = [
        c for c in summary_df.columns
        if isinstance(c, str) and re.match(r"^W-\d{1,2}$", c.strip())
    ]
    if not week and week_cols:
        week = sorted(week_cols, key=lambda x: int(x.split("-")[1]))[-1]

    zone = extract_zone(question, zones)
    if not zone:
        zone = "ALL INDIA" if "ALL INDIA" in zones else (zones[0] if zones else None)

    if not week or week not in summary_df.columns:
        return week, zone, []

    rows = select_relevant_rows(question, summary_df)

    result = []
    for r in rows:
        match = summary_df[summary_df[metric_col] == r]
        if not match.empty:
            val = match.iloc[0][week]
            result.append((str(r), "" if pd.isna(val) else str(val)))

        # expand header blocks
        if r in ("BUSINESS TYPE BREAKDOWN", "CN Status Breakdown"):
            start_idx = match.index[0] if not match.empty else None
            if start_idx is not None:
                for i in range(start_idx + 1, min(start_idx + 20, len(summary_df))):
                    mname = str(summary_df.iloc[i][metric_col])
                    if (
                        mname in ("BUSINESS TYPE BREAKDOWN", "CN Status Breakdown")
                        or mname.startswith("TPTR Mode")
                        or mname.startswith("Picked Vol. Zone")
                    ):
                        break
                    v = summary_df.iloc[i][week]
                    result.append((mname, "" if pd.isna(v) else str(v)))

    return week, zone, result


# =============================================================================
# ✅ STREAMLIT UI
# =============================================================================

st.set_page_config(page_title="EDD MIS Chatbot", layout="wide")

st.title("📦 EDD MIS Chatbot")
st.caption("Upload Excel → (1) View Summary  (2) Ask Questions  (3) Process Unprocessed File")

with st.sidebar:
    st.header("Settings")

    model = st.selectbox(
        "Model",
        options=["gpt-4o-mini", "gpt-4o"],
        index=0,
    )

    st.markdown("### 🔑 OpenAI Secret Key")
    st.caption("Preferred: `.streamlit/secrets.toml` → `OPENAI_API_KEY = \"sk-...\"`")
    key_in = st.text_input("Paste key (sk-...)", type="password")
    if key_in:
        st.session_state["_openai_key"] = key_in

uploaded = st.file_uploader("Upload EDD Excel (.xlsx)", type=["xlsx"], key="main_upload")

tabs = st.tabs(["📊 Summary View", "💬 Ask a Question", "🛠️ Process File"])

with tabs[0]:
    if not uploaded:
        st.info("Upload a processed Excel first (jisme summary sheet already ho).")
    else:
        xbytes = uploaded.getvalue()
        try:
            sdf = load_summary_df(xbytes, sheet="summary")
            st.subheader("Summary Sheet")
            st.dataframe(sdf, use_container_width=True, height=600)
        except Exception as e:
            st.error(f"Summary sheet read nahi ho paayi: {e}")

with tabs[1]:
    if not uploaded:
        st.info("Pehle Excel upload karo.")
    else:
        xbytes = uploaded.getvalue()
        try:
            summary_df = load_summary_df(xbytes, sheet="summary")
        except Exception as e:
            st.error(f"Summary sheet read nahi ho paayi: {e}")
            summary_df = pd.DataFrame()

        st.subheader("Ask Questions (Anything from Summary)")
        q = st.text_input(
            "Type your question:",
            placeholder="e.g. week-48 picked up volume / Week 51 ALL INDIA CN current status / week 46 ndr not available",
        )
        ask = st.button("Ask")

        if ask and q.strip():
            week, zone, items = answer_from_summary(summary_df, q)

            if not items:
                st.error("Week/metric match nahi hua. Example: 'week-48 picked volume', 'Week 51 ALL INDIA CN current status'")
            else:
                client = make_client()
                context_lines = "\n".join([f"- {m}: {v}" for m, v in items])
                prompt = f"""Question: {q}

Excel Summary for {week} (zone={zone}):
{context_lines}

Give a short, direct answer. Do NOT show JSON or code. If multiple rows, format as bullet points."""
                if client:
                    with st.spinner("Generating answer..."):
                        ans = llm_answer(
                            client=client,
                            model=model,
                            system="You are a logistics MIS analyst. Answer only from the provided Excel summary context. If info missing, say what is missing.",
                            user=prompt,
                        )
                    if ans.strip().startswith("❌"):
                        st.error(ans)
                    else:
                        st.success(ans)
                else:
                    st.success("\n".join([f"• {m}: {v}" for m, v in items]))

with tabs[2]:
    st.subheader("Process Unprocessed File → Generate Summary")
    st.write("Yahan tum apna backend processing (pandas+openpyxl) run kara sakte ho.")
    unp = st.file_uploader("Upload Unprocessed Excel (.xlsx)", type=["xlsx"], key="unprocessed_upload")

    if unp:
        if st.button("Run Processing"):
            with st.spinner("Processing... (summary + grouping + formatting)"):
                with tempfile.TemporaryDirectory() as td:
                    in_path = os.path.join(td, "input.xlsx")
                    out_path = os.path.join(td, "EDD_Report_Processed.xlsx")
                    with open(in_path, "wb") as f:
                        f.write(unp.getvalue())

                    try:
                        process_edd_report(
                            input_path=in_path,
                            output_path=out_path,
                            source_sheet="Query result",
                            summary_sheet="summary",
                            apply_formatting=True,
                        )

                        with open(out_path, "rb") as f:
                            out_bytes = f.read()

                        st.success("✅ Processing done! Download processed file below.")
                        st.download_button(
                            "⬇️ Download Processed Excel",
                            data=out_bytes,
                            file_name="EDD_Report_Processed.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except Exception as e:
                        st.error(f"Processing failed: {e}")
